from random import choice, randint
from string import ascii_letters
from os import path, listdir
from pathlib import Path
import openpyxl


def number_input() -> int:
    """Принимает ввод от пользователя, проверяет, число ли это, и возвращает его"""
    selected: bool = False
    while not selected:
        try:
            selection: int = int(input())
            selected = True
        except (ValueError):
            print('Неверный выбор.')
    return selection


def random_password() -> str:
    random_list = ascii_letters + "_0123456789"
    result = ''
    for i in range(0, randint(15, 30) + 1, 1):
        result += choice(random_list)
    return result


def read_from_xsl() -> list:
    print('Выберите файл (должен находиться в папке files):')
    Path('files').mkdir(parents=True, exist_ok=True)
    for number, file in enumerate(listdir('files')):
        print(f'{number}. {file}')
    file = openpyxl.load_workbook(f"files/{listdir('files')[number_input()]}")
    xl_list = file['Список пользователей']
    users: list[dict(str, str)] = []
    for row in xl_list.iter_rows(2, xl_list.max_row):
        if row[0].value is not None:
            user = {
                'name': f'{row[0].value} {row[1].value}',
                'email': row[2].value,
                'role': row[3].value,
            }
        else:
            break
        users.append(user)

    return users
